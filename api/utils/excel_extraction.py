import os
import pandas as pd
import numpy as np
import logging
import openpyxl
import xlrd
import re
import os
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font
from openpyxl.utils import get_column_letter

# Configure logging
logging.basicConfig(level=logging.DEBUG, format='%(asctime)s - %(levelname)s - %(message)s')


def copy_cell_format(source_cell, target_cell):
    """Copy all formatting from source cell to target cell."""
    if source_cell.has_style:
        target_cell.font = Font(
            name=source_cell.font.name,
            size=source_cell.font.size,
            bold=source_cell.font.bold,
            italic=source_cell.font.italic,
            vertAlign=source_cell.font.vertAlign,
            underline=source_cell.font.underline,
            strike=source_cell.font.strike,
            color=source_cell.font.color
        )
        
        target_cell.border = Border(
            left=Side(border_style=source_cell.border.left.style,
                     color=source_cell.border.left.color),
            right=Side(border_style=source_cell.border.right.style,
                      color=source_cell.border.right.color),
            top=Side(border_style=source_cell.border.top.style,
                    color=source_cell.border.top.color),
            bottom=Side(border_style=source_cell.border.bottom.style,
                       color=source_cell.border.bottom.color)
        )
        
        target_cell.fill = PatternFill(
            fill_type=source_cell.fill.fill_type,
            start_color=source_cell.fill.start_color,
            end_color=source_cell.fill.end_color
        )
        
        target_cell.alignment = Alignment(
            horizontal=source_cell.alignment.horizontal,
            vertical=source_cell.alignment.vertical,
            text_rotation=source_cell.alignment.text_rotation,
            wrap_text=source_cell.alignment.wrap_text,
            shrink_to_fit=source_cell.alignment.shrink_to_fit,
            indent=source_cell.alignment.indent
        )


def remove_rows_after_index(input_file, output_file, index_row):
    """
    Read Excel file (both .xls and .xlsx formats), remove all rows after the specified index,
    and save to a new file.
    
    Args:
        input_file: Path to input Excel file (.xls or .xlsx)
        output_file: Path to save the modified Excel file (will be saved as .xlsx)
        index_row: Index after which all rows should be removed (1-based index)
    """
    logging.info(f"Reading Excel file: {input_file}")
    logging.info(f"Index row to keep: {index_row}")
    
    # Check file extension
    file_ext = os.path.splitext(input_file)[1].lower()
    if file_ext == '.xls':
        logging.error("Only .xlsx files are supported currently")
        raise ValueError("Only .xlsx files are supported currently. Please convert your file to .xlsx format.")
    temp_file = None
    try:
        wb = openpyxl.load_workbook(input_file)
        # Process each sheet
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            max_row = ws.max_row
            if max_row > index_row:
                # Delete rows from bottom to top to avoid index shifting
                logging.info(f"Removing rows from {index_row + 1} to {max_row}")
                ws.delete_rows(index_row + 1, max_row - index_row)
        logging.info(f"Saving modified Excel file to: {output_file}")
        wb.save(output_file)
        logging.info("Completed successfully")
    finally:
        # Clean up temporary file if it was created
        if temp_file and os.path.exists(temp_file):
            try:
                os.remove(temp_file)
                logging.info("Cleaned up temporary file")
            except Exception as e:
                logging.warning(f"Failed to clean up temporary file: {str(e)}")



def extract_excel_data(file_path):
    logging.info(f"Starting extraction of data from Excel file: {file_path}")

    # Read the entire sheet without headers using a compatible engine
    df = _safe_read_excel(file_path, header=None)
    # Find the row containing 'Print No' using regex
    header_row_idx = None
    for idx, row in df.iterrows():
        if any(re.match(r'^print\s*no\.?\s*$', str(cell).strip().lower()) for cell in row):
            header_row_idx = idx
            logging.info(f"Header row index found at: {header_row_idx}")
            break
    if header_row_idx is None:
        raise KeyError("Could not find a row containing 'Print No'.")

    # Extract parent and sub-columns
    header_file_path = "temp_modified.xlsx"
    
    try:
        remove_rows_after_index(
            input_file=file_path,
            output_file=header_file_path,
            index_row=header_row_idx
        )
    except Exception as e:
        logging.error(f"Error processing Excel file: {str(e)}")
        raise  # Re-raise the exception to see the full error details
   
    parent_columns = df.iloc[header_row_idx]
    sub_columns = df.iloc[header_row_idx + 1]

    logging.debug(f"Parent columns: {parent_columns.tolist()}")
    logging.debug(f"Sub columns: {sub_columns.tolist()}")
    # Check if sub_columns has any non-empty data
    # Check if sub_columns contains 'Min' or 'Max' (case-insensitive, ignore NaN)
    has_sub_data = any(
        isinstance(cell, str) and ("min" in cell.lower() or "max" in cell.lower())
        for cell in sub_columns if not pd.isna(cell)
    )
    logging.info(f"###############has_sub_data: {has_sub_data}")
    if has_sub_data:
        # If no sub-column data, use parent columns directly
        combined_columns = []
        for parent, sub in zip(parent_columns, sub_columns):
            parent_str = str(parent).strip() if not pd.isna(parent) else ''
            sub_str = str(sub).strip() if not pd.isna(sub) else ''
            if parent_str and sub_str:
                combined_columns.append(f"{parent_str} {sub_str}")
            elif parent_str:
                combined_columns.append(parent_str)
            elif sub_str:
                combined_columns.append(sub_str)
            else:
                combined_columns.append('nan')
        logging.info("Found sub-column data, combining parent and sub columns")
        start_row = header_row_idx +2

    else:
        # Combine parent and sub-columns

        combined_columns = [
            str(col).strip() if not pd.isna(col) else 'nan'
            for col in parent_columns
        ]
        logging.info("No sub-column data found, using parent columns only")
        start_row = header_row_idx +1

        

    logging.debug(f"Combined columns before renaming: {combined_columns}")
    # Rename 'Min' to 'Tolerance Min'
    combined_columns = ['TOLERANCE MIN' if col.lower() == 'min' else col for col in combined_columns]
    combined_columns = ['TOLERANCE MAX' if col.lower() == 'max' else col for col in combined_columns]

    df.columns = combined_columns
    logging.debug(f"Columns after renaming: {df.columns.tolist()}")

    # Extract rows above header_row_idx while preserving the exact format
    # pre_header_df = df.iloc[:header_row_idx].copy()
    # pre_header_df.reset_index(drop=True, inplace=True)
    # logging.debug("Extracted pre-header data with exact format:")
    # logging.debug(pre_header_df)

    # Drop header rows
    df = df.drop(range(start_row))
    df = df.reset_index(drop=True)

    # Convert 'Print No' as key (generic, case-insensitive)
    # Find the column that contains both 'print' and 'no' (case-insensitive)
    key_col = None
    for col in df.columns:
        col_str = str(col).lower().replace(' ', '')
        if 'print' in col_str and 'no' in col_str:
            key_col = col
            break
    if key_col is None:
        raise KeyError("Could not find a column containing both 'print' and 'no'.")

    data_dict = {}
    for _, row in df.iterrows():
        key = row[key_col]
        value = row.to_dict()
        data_dict[key] = value
        #logging.debug(f"Extracted data for {key_col} {key}: {value}")

    logging.info(f"Extraction completed. Total items extracted: {len(data_dict)}")
    return data_dict, header_file_path,header_row_idx


def _safe_read_excel(file_path, **kwargs):
    """Read Excel using an explicit engine based on file extension.

    .xlsx -> openpyxl
    .xls  -> xlrd
    If engine is unavailable, raises ImportError with actionable message.
    """
    ext = os.path.splitext(file_path)[1].lower()
    engine = None
    if ext == '.xls':
        engine = 'xlrd'
    elif ext in ('.xlsx', '.xlsm', '.xltx', '.xltm'):
        engine = 'openpyxl'
    # If extension unknown, let pandas try but prefer openpyxl
    try:
        if engine:
            return pd.read_excel(file_path, engine=engine, **kwargs)
        return pd.read_excel(file_path, **kwargs)
    except ImportError as e:
        # Provide a helpful message about installing the required package
        if engine == 'xlrd':
            raise ImportError("xlrd is required to read .xls files. Install it with: pip install xlrd==1.2.0") from e
        if engine == 'openpyxl':
            raise ImportError("openpyxl is required to read .xlsx files. Install it with: pip install openpyxl") from e
        raise

if __name__ == "__main__":
    file_path = "final_inscpection.xlsx"  # Path to Excel file
    pre_header_df, extracted_data = extract_excel_data(file_path)
    print("Pre-header DataFrame:")
    print(pre_header_df)
    print("Extracted Data:")
    for print_no, data in extracted_data.items():
        print(f"Print No: {print_no}")
        print(data)