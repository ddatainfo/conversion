import pandas as pd
import numpy as np
import sys
import re
import logging
import openpyxl
from openpyxl.utils import get_column_letter
sys.path.append("../")  # Add parent directory to sys.path for relative imports
from api.utils.excel_extraction import extract_excel_data, copy_cell_format
from api.utils.extract_measurements import extract_measurements

# Configure logging
logging.basicConfig(level=logging.DEBUG, format='%(asctime)s - %(levelname)s - %(message)s')
COLUMN_TO_RMV = ['OUT OF TOLERANCE', 'DEVIATION', 'OUT_OF_TOLERANCE','IDENTIFICATION NO']  # replace with your list

# def merge_excel_with_header(output_file_path, header_file_path, final_output_path, header_row_idx):
#     """
#     Merge two Excel files while preserving the cell formatting from the header file.
    
#     Args:
#         output_file_path: Path to the Excel file containing the data
#         header_file_path: Path to the Excel file containing the header with formatting to preserve
#         final_output_path: Path where the merged file will be saved
#     """
#     logging.info("Starting Excel merge with format preservation")
#     logging.info(f"Data file: {output_file_path}, Header file: {header_file_path}")
    
#     try:
#         # Load both workbooks
#         header_wb = openpyxl.load_workbook(header_file_path)
#         data_wb = openpyxl.load_workbook(output_file_path)
        
#         # Create a new workbook and get the active sheet
#         merged_wb = openpyxl.Workbook()
#         merged_sheet = merged_wb.active
        
#         # If data workbook has 'combined' sheet, use that, otherwise use first sheet
#         data_sheet_name = 'combined' if 'combined' in data_wb.sheetnames else data_wb.sheetnames[0]
#         data_sheet = data_wb[data_sheet_name]
#         logging.info(f"Using data from sheet: {data_sheet_name}")
        
#         # Get the first sheet from header workbook
#         header_sheet_name = header_wb.sheetnames[0]
#         header_sheet = header_wb[header_sheet_name]
#         logging.info(f"Using header format from sheet: {header_sheet_name}")
        
#         # Set the sheet name
#         merged_sheet.title = header_sheet_name
        
#         # Copy column widths from header sheet
#         max_cols = max(header_sheet.max_column, data_sheet.max_column)
#         for col in range(1, max_cols + 1):
#             col_letter = get_column_letter(col)
#             if col_letter in header_sheet.column_dimensions:
#                 merged_sheet.column_dimensions[col_letter].width = header_sheet.column_dimensions[col_letter].width
#             else:
#                 merged_sheet.column_dimensions[col_letter].width = 10  # Default width
        
#         # Copy row heights from header sheet
#         for row in range(1, header_row_idx + 1):
#             if row in header_sheet.row_dimensions:
#                 merged_sheet.row_dimensions[row].height = header_sheet.row_dimensions[row].height
        
#         logging.info(f"header_sheet row:{header_row_idx}")
#         # First, copy header rows with formatting
#         for row in range(1, header_row_idx + 1):
#             for col in range(1, header_sheet.max_column + 1):
#                 source_cell = header_sheet.cell(row=row, column=col)
#                 target_cell = merged_sheet.cell(row=row, column=col)
                
#                 # Copy value and formatting
#                 target_cell.value = source_cell.value
#                 copy_cell_format(source_cell, target_cell)
        
#         # Then append data rows
#         start_row = header_row_idx + 1
#         for data_row in range(1, data_sheet.max_row + 1):
#             for col in range(1, data_sheet.max_column + 1):
#                 source_cell = data_sheet.cell(row=data_row, column=col)
#                 target_cell = merged_sheet.cell(row=start_row + data_row - 1, column=col)
#                 target_cell.value = source_cell.value
        
#         # Save the merged workbook
#         merged_wb.save(final_output_path)
#         logging.info(f"Successfully saved merged file to {final_output_path}")
        
#     except Exception as e:
#         logging.error(f"Error during Excel merge: {str(e)}")
#         raise
        
#         # Save the merged workbook
#         merged_wb.save(final_output_path)
#         logging.info(f"Successfully saved merged file to {final_output_path}")
        
#     except Exception as e:
#         logging.error(f"Error during Excel merge: {str(e)}")
#         raise
def merge_excel_with_header(output_file_path, header_file_path, final_output_path, header_row_idx):
    """
    Append data workbook into header workbook starting after header_row_idx,
    preserving header formatting from header_file_path.
    """
    logging.info("Starting Excel merge (append data into header) with format preservation")
    logging.info(f"Data file: {output_file_path}, Header file: {header_file_path}, header_row_idx: {header_row_idx}")

    try:
        # Load workbooks
        header_wb = openpyxl.load_workbook(header_file_path)
        data_wb = openpyxl.load_workbook(output_file_path)

        # Choose sheets (first sheet for header, 'combined' or first for data)
        header_sheet = header_wb[header_wb.sheetnames[0]]
        data_sheet_name = 'combined' if 'combined' in data_wb.sheetnames else data_wb.sheetnames[0]
        data_sheet = data_wb[data_sheet_name]

        logging.info(f"Using data from sheet: {data_sheet_name}")
        logging.info(f"Using header format from sheet: {header_wb.sheetnames[0]}")

        # Determine column range to handle
        max_cols = max(header_sheet.max_column, data_sheet.max_column)

        # Ensure column widths are preserved/extended
        for col in range(1, max_cols + 1):
            col_letter = get_column_letter(col)
            logging.info(f"Processing column: {col_letter}")
            if col_letter in header_sheet.column_dimensions and header_sheet.column_dimensions[col_letter].width:
                # keep existing width
                header_w = header_sheet.column_dimensions[col_letter].width
                header_sheet.column_dimensions[col_letter].width = header_w
            elif col_letter in data_sheet.column_dimensions and data_sheet.column_dimensions[col_letter].width:
                # fallback to data sheet width if header has none
                header_sheet.column_dimensions[col_letter].width = data_sheet.column_dimensions[col_letter].width
            else:
                # ensure a default width exists
                header_sheet.column_dimensions[col_letter].width = header_sheet.column_dimensions.get(col_letter, openpyxl.worksheet.dimensions.ColumnDimension(header_sheet, index=col_letter)).width or 10

        # Insert empty rows after header_row_idx to accommodate all data rows
        data_rows = data_sheet.max_row
        if data_rows > 0:
            insert_at = header_row_idx + 1
            header_sheet.insert_rows(insert_at, amount=data_rows)

            # For each data row, copy values and apply formatting.
            # Prefer copying formatting from the header_row_idx row (header template) if available,
            # otherwise copy cell formatting from source data cell.
            for r in range(1, data_rows + 1):
                target_row = insert_at + r - 1
                for c in range(1, max_cols + 1):
                    source_cell = data_sheet.cell(row=r, column=c) if c <= data_sheet.max_column else None
                    target_cell = header_sheet.cell(row=target_row, column=c)
                    # copy value
                    if source_cell is not None:
                        target_cell.value = source_cell.value
                    else:
                        target_cell.value = None

                    # Prefer header template formatting from the last header row
                    header_template_cell = header_sheet.cell(row=header_row_idx, column=c) if c <= header_sheet.max_column else None
                    try:
                        if header_template_cell is not None and header_template_cell.has_style:
                            copy_cell_format(header_template_cell, target_cell)
                        elif source_cell is not None and getattr(source_cell, "has_style", False):
                            copy_cell_format(source_cell, target_cell)
                    except Exception:
                        # protect against any unexpected style-copy issues per cell
                        logging.debug(f"Failed to copy style for row {target_row} col {c}", exc_info=True)
                        continue

        # Save result back to final_output_path (overwrite or new file)
        header_wb.save(final_output_path)
        logging.info(f"Successfully saved merged file to {final_output_path}")

    except Exception as e:
        logging.error(f"Error during Excel merge: {str(e)}", exc_info=True)
        raise
def move_measured_columns_to_end(df):
    """
    Reorder DataFrame columns so all columns starting with 'MEASURED' (case-insensitive)
    appear at the end, preserving the order of other columns.
    Returns a new DataFrame with reordered columns.
    """
    cols = df.columns.tolist()
    measured_cols = [col for col in cols if str(col).strip().upper().startswith("MEASURED")]
    other_cols = [col for col in cols if not str(col).strip().upper().startswith("MEASURED")]
    return df[other_cols + measured_cols]

def _get_merged_cell_value(sheet, row, col):
    """Get cell value, handling merged cells properly."""
    cell = sheet.cell(row=row, column=col)
    for range_ in sheet.merged_cells.ranges:
        if cell.coordinate in range_:
            # Return value from the top-left cell of the merged range
            return sheet.cell(row=range_.min_row, column=range_.min_col).value
    return cell.value

def get_data_sheet_columns(sheet, header_row=1):
    """Get column headers from sheet, handling merged cells and stripping whitespace."""
    max_col = sheet.max_column
    headers = []
    for col in range(1, max_col + 1):
        value = _get_merged_cell_value(sheet, header_row, col)
        if value is not None:
            # Strip whitespace and trailing dots
            value = str(value).strip().rstrip('.')
        headers.append(value)
    return headers

def _try_float(v):
    """Safely convert v to float; return np.nan on failure."""
    try:
        if v is None:
            return np.nan
        return float(v)
    except Exception:
        return np.nan

def final_data(excel_file_path, txt_file_paths, output_file_path):
    """Merge Excel templates with one or more TXT measurement files.

    txt_file_paths may be a single path (str) or a list of paths. When multiple TXT files
    are provided, measured values are written into columns named MEASURED-1, MEASURED-2, ...
    """
    logging.info("Starting data merging process.")

    # Extract data from Excel file
    excel_data, header_file_path,header_row_idx = extract_excel_data(excel_file_path)
    logging.debug(f"excel_Data keys: {list(excel_data)}")

    # Normalize keys inside excel_data templates to uppercase so they match pre_header columns
    for k, v in list(excel_data.items()):
        excel_data[k] = { (col.upper() if isinstance(col, str) else col): val for col, val in v.items() }

    #logging.debug(f"Normalized excel_Data keys: {excel_data.values()}")
    # # Normalize pre_header column names to uppercase for matching
    # pre_header.columns = [col.upper() for col in pre_header.columns]
    # pre_header = pre_header.reset_index(drop=True)

    # Accept either a single path or a list of paths
    if isinstance(txt_file_paths, (str, bytes)):
        txt_file_paths = [txt_file_paths]

    # For each TXT file, extract measurements and build a mapping dim->first_measurement
    per_file_maps = []  # list of dicts: [{dim: measurement, ...}, ...]
    for txt_path in txt_file_paths:
        file_meas = extract_measurements(txt_path)
        mmap = {}
        for mes in file_meas:
            if '#' in mes.get('dimension', ''):
                try:
                    dp = mes.get('dimension', '').split('=')[0]
                    d = re.search(r'#(\d+)', dp)
                    if d:
                        dn = int(d.group(1))
                        # keep first measurement for this dimension in this file
                        if dn not in mmap:
                            mmap[dn] = mes
                except Exception:
                    continue
        per_file_maps.append(mmap)

    logging.debug(f"Per-file measurement maps count: {len(per_file_maps)}")

    # Build merged_data by iterating excel_data templates; create MEASURED-N columns for each file
    merged_data = []
    unmatched_data = []

    multi_files = len(per_file_maps) > 1

    for key, template in excel_data.items():
        base = template.copy()
        # For multiple files, add MEASURED-1..N; for single file, use 'MEASURED'
        if multi_files:
            for idx, mmap in enumerate(per_file_maps, start=1):
                mes = mmap.get(key)
                colname = f"MEASURED-{idx}"
                if mes is not None:
                    base[colname] = _try_float(mes.get('measured'))
                else:
                    base[colname] = np.nan
            # keep original DEVIATION/OUT OF TOLERANCE empty (or could compute from first file)
            merged_data.append(base)
        else:
            # single file behavior: populate MEASURED, DEVIATION, OUT OF TOLERANCE if available
            mmap = per_file_maps[0] if per_file_maps else {}
            mes = mmap.get(key)
            if mes is not None:
                if mes.get('+tol') is not None:
                    base['TOLERANCE MAX'] = _try_float(mes.get('+tol'))
                if mes.get('-tol') is not None:
                    base['TOLERANCE MIN'] = _try_float(mes.get('-tol'))
                
                base['DEVIATION'] = _try_float(mes.get('deviation'))
                base['OUT OF TOLERANCE'] = _try_float(mes.get('outtol'))
                base['MEASURED'] = _try_float(mes.get('measured'))
            else:
                # ensure MEASURED exists
                base.setdefault('MEASURED', '')
            merged_data.append(base)

    # Any measurement keys not present in excel_data are unmatched
    logging.debug(f"MEger_data count before unmatched: {merged_data}")
    all_keys_in_files = set().union(*[set(m.keys()) for m in per_file_maps]) if per_file_maps else set()
    logging.info(f"All keys in measurement files: {all_keys_in_files}")
    unmatched_keys = all_keys_in_files - set(excel_data.keys())
    logging.info(f"Unmatched keys count: {unmatched_keys}")
    for uk in unmatched_keys:
        for mmap in per_file_maps:
            mes = mmap.get(uk)
            if mes:
                unmatched_record = {
                    'DIMENSION_NUMBER': uk,
                    'DIMENSION': mes.get('dimension'),
                    'TOLERANCE_MAX': mes.get('+tol'),
                    'TOLERANCE_MIN': mes.get('-tol'),
                    'DEVIATION': mes.get('deviation'),
                    'OUT_OF_TOLERANCE': mes.get('outtol'),
                    'MEASURED': mes.get('measured')
                }
                unmatched_data.append(unmatched_record)

    logging.debug(f"unmatxhed_data count: {unmatched_data}")
    logging.info(f"Built merged_data rows: {len(merged_data)}; unmatched: {len(unmatched_data)}")

    # Convert merged_data to DataFrame and drop columns that are all NaN
    merged_df = pd.DataFrame(merged_data)
    merged_df = merged_df.dropna(axis=1, how='all')
    logging.debug(f"Merged DataFrame columns before dropping specified columns: {merged_df.columns.tolist()}")
    cols_to_rmv = [c for c in COLUMN_TO_RMV if c in merged_df.columns]
    logging.debug(f"Columns to be removed: {cols_to_rmv}")
    if cols_to_rmv:
       merged_df.drop(columns=cols_to_rmv, inplace=True)
    merged_df = move_measured_columns_to_end(merged_df)
    logging.debug(f"Merged DataFrame columns after dropping all-NaN and reordering: {merged_df.columns.tolist()}")
    # Save the data to a temporary Excel file first
    # temp_output = output_file_path
    temp_output = output_file_path
    logging.info(f"Writing to temporary file: {temp_output}")
    logging.debug(f"Merged DataFrame preview:\n{merged_df.head()}")
    merged_df.to_excel("data.xlsx", index=False)
    # with pd.ExcelWriter(temp_output, engine='openpyxl') as writer:
    #     logging.info(f"Writing merged data to temporary file: {temp_output}")
    #     merged_df.to_excel(writer, sheet_name='Sheet 1', index=False)
    #     if unmatched_data:
    #         logging.info(f"Writing unmatched data to temporary file: {temp_output}")
    #         pd.DataFrame(unmatched_data).to_excel(writer, sheet_name='unmatched', index=False)
    
    # Merge the temporary file with the header file while preserving formatting
    try:
        logging.info("Merging temporary file with header file to preserve formatting.")
        merge_excel_with_header("data.xlsx", header_file_path, temp_output,header_row_idx)
        logging.info(f"Final formatted data saved to {temp_output}")
    
    except Exception as e:
        logging.error(f"Failed to create excel file: {str(e)}")


